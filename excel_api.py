from datetime import datetime
from collections import namedtuple

from openpyxl.cell import Cell
from openpyxl import load_workbook


DisciplineType = namedtuple('Discipline', 'number name study_group lecturer room')


class ExcelApi:
    def __init__(self, filename):
        self.wb = load_workbook(filename, data_only=True)
        self.sheet = self.get_special_sheet(self.wb)

        if self.sheet is not None:
            self.disciplines = self.get_dates_with_disciplines(self.sheet)
        else:
            self.disciplines = {}

    @staticmethod
    def get_special_sheet(workbook):
        for sheet_name in workbook.sheetnames:
            sheet = workbook.get_sheet_by_name(sheet_name)

            if sheet.cell(row=1, column=1).value is not None:
                for i in range(1, 1200):
                    if isinstance(sheet.cell(row=i, column=1).value, datetime):
                        return sheet

    @staticmethod
    def get_dates_with_disciplines(sheet) -> dict:
        disciplines = {}
        for i in range(1, 1200):
            cell = sheet.cell(row=i, column=1)

            if isinstance(cell.value, datetime):
                disciplines[cell.value.date()] = {
                    'cell': cell,
                    'disciplines': ExcelApi.add_disciplines(sheet, cell),
                }
        return disciplines

    @staticmethod
    def add_discipline(sheet, disciplines, number, row, col):
        if sheet.cell(row=row, column=col).value is None:
            return

        name = sheet.cell(row=row, column=col).value
        study_group = sheet.cell(row=4, column=col).value

        lecturer1 = sheet.cell(row=row, column=col + 5).value
        if lecturer1 is not None:
            room_value = sheet.cell(row=row, column=col + 7).value
            room = room_value if room_value is not None else sheet.cell(row=4, column=col + 7).value
            disciplines.append(DisciplineType(number, name, study_group, lecturer1, room))

        lecturer2 = sheet.cell(row=row, column=col + 6).value
        if lecturer2 is not None:
            room_value = sheet.cell(row=row, column=col + 8).value
            room = room_value if room_value is not None else sheet.cell(row=4, column=col + 8).value
            disciplines.append(DisciplineType(number, name, study_group, lecturer2, room))

    @staticmethod
    def add_disciplines(sheet, date_cell: Cell) -> list:
        row, col = date_cell.row, date_cell.column + 4

        disciplines = []
        while col < 300:
            ExcelApi.add_discipline(sheet, disciplines, '1-2', row, col)
            ExcelApi.add_discipline(sheet, disciplines, '3-4', row + 1, col)
            ExcelApi.add_discipline(sheet, disciplines, '5-6', row + 2, col)
            ExcelApi.add_discipline(sheet, disciplines, '7-8', row + 3, col)
            col += 9

        return disciplines

    def get_discipline(self, date, lecturer) -> list:
        res_disciplines = []

        if self.disciplines.get(date) is None:
            return res_disciplines

        for discipline in self.disciplines[date]['disciplines']:
            if discipline.lecturer == lecturer:
                res_disciplines.append(discipline)
        return res_disciplines
